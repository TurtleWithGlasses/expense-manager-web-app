"""
Excel Export Service for generating detailed financial reports
"""
import io
from datetime import date, datetime
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.entry import Entry
from app.models.category import Category
from app.core.currency import CurrencyService
from app.services.user_preferences import UserPreferencesService


class ExcelExportService:
    def __init__(self):
        self.currency_service = CurrencyService()
        self.user_preferences_service = UserPreferencesService()
    
    async def export_entries_to_excel(
        self,
        db: Session,
        user_id: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        category_id: Optional[int] = None,
        report_type: str = "all"
    ) -> io.BytesIO:
        """
        Export entries to Excel format with detailed information
        
        Args:
            db: Database session
            user_id: User ID to filter entries
            start_date: Start date for filtering (optional)
            end_date: End date for filtering (optional)
            category_id: Category ID for filtering (optional)
            report_type: Type of report ("all", "income", "expense")
        
        Returns:
            BytesIO object containing the Excel file
        """
        # Get user currency
        user_currency = self.user_preferences_service.get_user_currency(db, user_id)
        
        # Build query
        query = db.query(Entry).filter(Entry.user_id == user_id)
        
        # Apply date filters
        if start_date:
            query = query.filter(Entry.date >= start_date)
        if end_date:
            query = query.filter(Entry.date <= end_date)
        
        # Apply category filter
        if category_id:
            query = query.filter(Entry.category_id == category_id)
        
        # Apply type filter
        if report_type in ["income", "expense"]:
            query = query.filter(Entry.type == report_type)
        
        # Get entries with category information
        entries = query.join(Category, Entry.category_id == Category.id).order_by(Entry.date.desc()).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add title and metadata
        ws['A1'] = "Financial Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        
        # Add report information
        row = 3
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1
        
        ws[f'A{row}'] = "Date Range:"
        if start_date and end_date:
            ws[f'B{row}'] = f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        elif start_date:
            ws[f'B{row}'] = f"From {start_date.strftime('%Y-%m-%d')}"
        elif end_date:
            ws[f'B{row}'] = f"Until {end_date.strftime('%Y-%m-%d')}"
        else:
            ws[f'B{row}'] = "All time"
        row += 1
        
        ws[f'A{row}'] = "Category:"
        if category_id:
            category = db.query(Category).filter(Category.id == category_id).first()
            ws[f'B{row}'] = category.name if category else "Unknown"
        else:
            ws[f'B{row}'] = "All Categories"
        row += 1
        
        ws[f'A{row}'] = "Type:"
        ws[f'B{row}'] = report_type.title() if report_type != "all" else "All Types"
        row += 1
        
        ws[f'A{row}'] = "Currency:"
        ws[f'B{row}'] = user_currency
        row += 2
        
        # Add headers
        headers = [
            "Date", "Type", "Category", "Description", 
            "Amount (Original)", "Currency", "Amount (Converted)", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        row += 1
        
        # Add data rows
        total_income = 0
        total_expense = 0
        
        for entry in entries:
            # Convert amount to user currency
            converted_amount = await self.currency_service.convert_amount(
                entry.amount, entry.currency, user_currency
            )
            
            # Add row data
            ws.cell(row=row, column=1, value=entry.date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=entry.type.title())
            ws.cell(row=row, column=3, value=entry.category.name)
            ws.cell(row=row, column=4, value=entry.description)
            ws.cell(row=row, column=5, value=entry.amount)
            ws.cell(row=row, column=6, value=entry.currency)
            ws.cell(row=row, column=7, value=round(converted_amount, 2))
            ws.cell(row=row, column=8, value=entry.notes or "")
            
            # Apply borders to all cells in the row
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
            
            # Sum totals
            if entry.type.lower() == "income":
                total_income += converted_amount
            else:
                total_expense += converted_amount
            
            row += 1
        
        # Add summary section
        row += 1
        ws.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True, size=14)
        row += 1
        
        ws.cell(row=row, column=1, value="Total Income:")
        ws.cell(row=row, column=2, value=round(total_income, 2))
        ws.cell(row=row, column=3, value=user_currency)
        row += 1
        
        ws.cell(row=row, column=1, value="Total Expense:")
        ws.cell(row=row, column=2, value=round(total_expense, 2))
        ws.cell(row=row, column=3, value=user_currency)
        row += 1
        
        ws.cell(row=row, column=1, value="Net Balance:")
        ws.cell(row=row, column=2, value=round(total_income - total_expense, 2))
        ws.cell(row=row, column=3, value=user_currency)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    async def export_category_summary_to_excel(
        self,
        db: Session,
        user_id: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> io.BytesIO:
        """
        Export category-wise summary to Excel
        
        Args:
            db: Database session
            user_id: User ID to filter entries
            start_date: Start date for filtering (optional)
            end_date: End date for filtering (optional)
        
        Returns:
            BytesIO object containing the Excel file
        """
        # Get user currency
        user_currency = self.user_preferences_service.get_user_currency(db, user_id)
        
        # Build query
        query = db.query(Entry).filter(Entry.user_id == user_id)
        
        # Apply date filters
        if start_date:
            query = query.filter(Entry.date >= start_date)
        if end_date:
            query = query.filter(Entry.date <= end_date)
        
        # Get entries with category information
        entries = query.join(Category, Entry.category_id == Category.id).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Category Summary"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws['A1'] = "Category Summary Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Add report information
        row = 3
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1
        
        ws[f'A{row}'] = "Date Range:"
        if start_date and end_date:
            ws[f'B{row}'] = f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        elif start_date:
            ws[f'B{row}'] = f"From {start_date.strftime('%Y-%m-%d')}"
        elif end_date:
            ws[f'B{row}'] = f"Until {end_date.strftime('%Y-%m-%d')}"
        else:
            ws[f'B{row}'] = "All time"
        row += 1
        
        ws[f'A{row}'] = "Currency:"
        ws[f'B{row}'] = user_currency
        row += 2
        
        # Add headers
        headers = ["Category", "Income", "Expense", "Net"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row += 1
        
        # Calculate category totals
        category_totals = {}
        for entry in entries:
            category_name = entry.category.name
            if category_name not in category_totals:
                category_totals[category_name] = {"income": 0, "expense": 0}
            
            converted_amount = await self.currency_service.convert_amount(
                entry.amount, entry.currency, user_currency
            )
            
            if entry.type.lower() == "income":
                category_totals[category_name]["income"] += converted_amount
            else:
                category_totals[category_name]["expense"] += converted_amount
        
        # Add category data
        for category_name, totals in sorted(category_totals.items()):
            ws.cell(row=row, column=1, value=category_name)
            ws.cell(row=row, column=2, value=round(totals["income"], 2))
            ws.cell(row=row, column=3, value=round(totals["expense"], 2))
            ws.cell(row=row, column=4, value=round(totals["income"] - totals["expense"], 2))
            
            # Apply borders
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
